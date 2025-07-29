import pyautogui
import pyperclip
from openpyxl import load_workbook
from openpyxl.styles import PatternFill
import time
import os
import sys
import json

# === FUNÇÃO PARA CARREGAR CONFIGURAÇÃO ===
def carregar_config():
    if getattr(sys, 'frozen', False):
        # Se estiver rodando como .exe
        pasta_base = os.path.dirname(sys.executable)
    else:
        # Rodando como script python
        pasta_base = os.path.dirname(os.path.abspath(__file__))
    caminho_config = os.path.join(pasta_base, "config.json")

    with open(caminho_config, encoding="utf-8") as f:
        config = json.load(f)
    return config, pasta_base

# === CARREGA CONFIGURAÇÃO ===
config, pasta_base = carregar_config()

# Usa os valores do config.json
caminho_excel = os.path.join(pasta_base, config["nome_planilha"])
coluna_chaves = config["coluna_chave"]
linha_inicial = config["linha_inicial"]
x_campo_chave = config["x_campo_chave"]
y_campo_chave = config["y_campo_chave"]
x_botao_salvar = config["x_botao_salvar"]
y_botao_salvar = config["y_botao_salvar"]
caminho_imagem_campo = os.path.join(pasta_base, config["imagem_chave"])
tempo_espera_antes_inicio = config.get("tempo_espera_antes_inicio", 15)
tempo_espera_entre_chaves = config.get("tempo_espera_entre_chaves", 5)

# === FUNÇÃO PARA EXECUTAR PROCESSO DE INSERÇÃO ===
def executar_processo():
    wb = load_workbook(caminho_excel)
    planilha = wb.active
    linha = linha_inicial

    print(f"O script começará em {tempo_espera_antes_inicio} segundos. Prepare a tela em modo 'tela cheia'...")
    time.sleep(tempo_espera_antes_inicio)

    while True:
        try:
            # Verificar se o campo está visível na tela
            if not pyautogui.locateOnScreen(caminho_imagem_campo, confidence=0.8):
                raise Exception("Campo de chave não encontrado")

            # Ler a próxima chave
            celula = f"{coluna_chaves}{linha}"
            chave = str(planilha[celula].value)

            if chave == "None" or chave.strip() == "":
                print("Fim das chaves ou célula vazia. Encerrando...")
                break

            # Copiar chave
            pyperclip.copy(chave)

            # Clicar e colar
            pyautogui.click(x=x_campo_chave, y=y_campo_chave)
            time.sleep(0.5)
            pyautogui.hotkey("ctrl", "a")
            pyautogui.press("delete")
            time.sleep(0.5)
            pyautogui.hotkey("ctrl", "v")
            print(f"Chave colada: {chave}")

            # Clicar em "Salvar Nota"
            time.sleep(1)
            pyautogui.click(x=x_botao_salvar, y=y_botao_salvar)
            print("Botão 'Salvar Nota' clicado.")

            # Pintar a célula de verde
            planilha[celula].fill = PatternFill(start_color="90EE90", end_color="90EE90", fill_type="solid")

            # Salvar progresso
            wb.save(caminho_excel)

            # Esperar antes da próxima
            time.sleep(tempo_espera_entre_chaves)
            linha += 1

        except Exception as e:
            print(f"\n⚠️ Erro detectado: {e}")
            print("Parece que a tela foi atualizada ou algo inesperado ocorreu.")

            while True:
                acao = input("\nDigite 1 para continuar após ajustar a tela, ou 2 para encerrar o script: ")
                if acao == '1':
                    print(f"Ok. O script reiniciará em {tempo_espera_antes_inicio} segundos...")
                    time.sleep(tempo_espera_antes_inicio)
                    executar_processo()  # Reinicia a função
                    return
                elif acao == '2':
                    print("Encerrando o script conforme solicitado.")
                    return
                else:
                    print("Entrada inválida. Digite apenas 1 ou 2.")

# === INICIAR ===
executar_processo()
